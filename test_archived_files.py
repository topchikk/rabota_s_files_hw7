
from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
import csv



from conftest import PATH_TO_ARCHIVE_DIR


def test_pdf():
    with ZipFile(PATH_TO_ARCHIVE_DIR+"/archived_file.zip") as zip_file:
        with zip_file.open("pdf123.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            assert "Содержание" in page.extract_text()


def test_xlsx():
    with ZipFile(PATH_TO_ARCHIVE_DIR+"/archived_file.zip") as zip_file:
        with zip_file.open("Budget_Template.xlsx") as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            assert sheet.cell(row=2,column=1).value == "Планируемый бюджет"


def test_csv():
    with ZipFile(PATH_TO_ARCHIVE_DIR+"/archived_file.zip") as zip_file:
        with zip_file.open("users.csv") as csv_file:
            content = csv_file.read().decode("utf-8-sig")
            reader = list(csv.reader(content.splitlines()))
            third_row = reader[2]
            assert ",".join(third_row[:]) == "Сергеев,Сергей,Сергеевич,Разработчик"
